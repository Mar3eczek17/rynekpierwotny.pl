"""
Aby uruchomić script, należy uzupełnić klucz API w pkt. 3 i wcisnąć przycisk "Run 'script'"
"""
"""
1.0 Napisz skrypt w Pythonie, który wczyta plik XLSX z danymi sprzedaży mieszkań z Warszawy i wyświetli na ekranie 10
pierwszych rekordów.
"""
import sqlite3
import pandas as pd
from KEY import API_KEY

# Wczytaj plik XLSX do dataframe
df = pd.read_excel('dane_do_zadania_.xlsx')

# Wyświetl 10 pierwszych rekordów
print(df.head(10))

"""
1.1 Preprocessing
"""
# Sprawdzenie, czy istnieją puste wartości w kolumnach
print(df.isnull().sum())

"""
Uzupełnienie brakujących wartości metodami: forward filling i backward filling, dla lepszego zoptymalizowania danych.
Stwierdzam, że dla obu kolumn (cena_m2 i cena), jest od 36,17% do 36,77%, pustych wartości. Dlatego przyjmuję optymalną
procedurę uzpupełninia pustych wartości, aby zróżnicować dane, biorąc pod uwagę, że dane są rozproszone.
"""
df['cena_m2'] = df['cena_m2'].fillna(method='bfill')
df['cena'] = df['cena'].fillna(method='ffill')

"""
2.0 W pliku XLSX znajdują się kolumny z informacjami o powierzchni i cenie mieszkań. Napisz skrypt w Pythonie, który
obliczy średnią cenę za metr kwadratowy w podziale na inwestycje w Warszawie.
"""

"""
Dane są sprzeczne! Jeśli cena = 784452 i powierzchnia = 48,5, to cena_m2 != 16600.
Powinno być == 16174,27 co daje w zaokrągleniu, cenę == 784452.
Dlatego przyjmuję, dwie opcje, jedną dla "cena_m2".
Drugą dla "cena" i "powierzchnia", z czego zliczam cena_m2_new.
Co więcej, aby pogrupować inwestycje, najbardziej logiczne grupowanie opiera się o kolumnę o nazwie: "lokalizacja" lub
"offer_id".
"""

# Opcja 1: Grupowanie danych według lokalizacji i obliczenie średniej ceny za metr kwadratowy
avg_price = df.groupby('lokalizacja')['cena_m2'].mean()

# Wyświetlenie wyniku
print(avg_price)

# Opcja 2: Dodanie kolumny z informacją o cenie za metr kwadratowy
df['cena_m2_new'] = df['cena'] / df['powierzchnia']

# Grupowanie danych po inwestycji i obliczenie średniej ceny za metr kwadratowy
avg_price = df.groupby('lokalizacja')['cena_m2_new'].mean()

# Wyświetlenie wyniku
print(avg_price)

# Sprawdzenie, grupowanie po "offer_id".
print("Opcja 1 sprawdzenie")
avg_price = df.groupby('offer_id')['cena_m2'].mean()

# Wyświetlenie wyniku
print(avg_price)

# Opcja 2: Dodanie kolumny z informacją o cenie za metr kwadratowy
print("Opcja 2 sprawdzenie")
df['cena_m2_new'] = df['cena'] / df['powierzchnia']

# Grupowanie danych po inwestycji i obliczenie średniej ceny za metr kwadratowy
avg_price = df.groupby('offer_id')['cena_m2_new'].mean()

# Wyświetlenie wyniku
print(avg_price)

# Spr. ogólne: 33 == "Sąsiedzka 6".
avg_price = df.loc[df['offer_id'] == 33, 'cena_m2_new'].mean()
avg_price_lokalizacja = df.loc[df['lokalizacja'] == "Sąsiedzka 6", 'cena_m2_new'].mean()

# Wyświetlenie wyniku
print(avg_price)
print(avg_price_lokalizacja)

# Spr. ogólne: 37 == "ul. Strumykowa 6".
avg_price = df.loc[df['offer_id'] == 37, 'cena_m2_new'].mean()
avg_price_lokalizacja = df.loc[df['lokalizacja'] == "ul. Strumykowa 6", 'cena_m2_new'].mean()

# Wyświetlenie wyniku
print(avg_price)
print(avg_price_lokalizacja)

"""
Jak widać dla niektórych "offer_id" są duże różnice dla niektórych małe.
"""

"""
3. W pliku XLSX znajduje się kolumna "lokalizacja", która zawiera adresy mieszkań. Napisz skrypt w Pythonie, który 
wykorzystując moduł Geopy i API Google Maps, pobierze współrzędne geograficzne (długość i szerokość geograficzną) dla 
każdego adresu i zapisze je do nowej kolumny "współrzędne" w pliku XLSX.
"""
import openpyxl
from geopy.geocoders import GoogleV3

# Tworzenie obiektu geokodera z kluczem API.
"""
W celu uzyskania klucza API Google Maps. 
Należy wejść na stronę internetową i zapoznać się z informacjami w niej zawartymi.
Poniżej załączam stronię i informacje na temat tego, jak uzyskać klucz API:
    Temat: Jak uzyskać klucz do API Google Maps i inne najczęściej zadawane pytania dotyczące map Google?
    Strona internetowa: https://www.idosell.com/pl/jak-uzyskac-klucz-do-api-google-map/
"""
geolocator = GoogleV3(api_key=API_KEY)

# Wczytywanie pliku XLSX
wb = openpyxl.load_workbook('dane_do_zadania_.xlsx')
sheet = wb.active

# Dodawanie nagłówka dla nowej kolumny
sheet.cell(row=1, column=12, value='współrzędne')

# Iteracja po wierszach i pobieranie współrzędnych geograficznych dla każdego adresu
for row in sheet.iter_rows(min_row=2, values_only=True):
    address = row[8]  # adres z kolumny "lokalizacja"
    location = geolocator.geocode(address)  # pobieranie współrzędnych geograficznych za pomocą Geopy i Google Maps API
    print(location)  # sprawdzenie

    if location is not None:
        latitude = location.latitude  # szerokość geograficzna
        longitude = location.longitude  # długość geograficzna
        coordinates = f"{latitude}, {longitude}"  # łączenie współrzędnych w jeden łańcuch znaków
        sheet.cell(row=row[0], column=12, value=coordinates)  # zapisywanie współrzędnych w nowej kolumnie
        print(latitude, longitude)  # sprawdzenie

# Zapisywanie zmian do pliku XLSX
wb.save('liczba_sprzedazy_mieszkan_na_miesiac.xlsx')

"""
4.0 Preprocessing
"""
df['data_sprzedazy'] = df['data_sprzedazy'].fillna(method='ffill')

"""
4.1 W pliku XLSX znajduje się kolumna "data_sprzedazy", która zawiera daty sprzedaży mieszkań.
Napisz skrypt w Pythonie, który za pomocą SQL stworzy tabelę w bazie danych zawierającą liczbę
sprzedaży mieszkań dla każdego miesiąca. Wyniki zapisz w nowym pliku XLSX.
"""
# Utwórz połączenie z bazą danych
conn = sqlite3.connect('baza_danych.db')

# Zapisz ramkę danych do tabeli w bazie danych
df.to_sql('sprzedaze', conn, if_exists='replace', index=False)

# Wykonaj zapytanie SQL i zapisz wyniki do nowej ramki danych
query = '''
        SELECT strftime('%Y-%m', data_sprzedazy) AS miesiac, COUNT(*) AS liczba_sprzedazy
        FROM sprzedaze
        GROUP BY strftime('%Y-%m', data_sprzedazy)
        '''
df2 = pd.read_sql_query(query, conn)

# Zapisz wyniki do nowego pliku XLSX
df2.to_excel('liczba_sprzedanych_mieszkan_na_miesiac.xlsx', index=False)

print(df2.head())

"""
5.0 W pliku XLSX znajduje się kolumna "data_sprzedazy", która zawiera daty sprzedaży mieszkań.
Napisz skrypt w Pythonie, który za pomocą SQL stworzy tabelę w bazie danych zawierającą liczbę
sprzedaży mieszkań dla każdego miesiąca. Wyniki zapisz w nowym pliku XLSX.
"""

"""
Jednym z sposobów na optymalizację przechowywania danych o sprzedaży mieszkań w bazie danych jest zastosowanie indeksów.
Indeksy pozwalają na szybszy dostęp do danych w tabeli poprzez utworzenie dodatkowej struktury danych, która zawiera 
klucze i odnośniki do wierszy w tabeli. Dzięki temu zapytania SQL, które wykorzystują kolumny, dla których utworzono 
indeksy, mogą być wykonane znacznie szybciej.

Kolejnym sposobem na optymalizację przechowywania danych jest zastosowanie partycjonowania. Partycjonowanie polega na 
podziale tabeli na mniejsze, bardziej zarządzalne części. Dzięki temu operacje takie jak wyszukiwanie, sortowanie 
i grupowanie danych można wykonywać tylko na wybranej części tabeli, co przyspiesza przetwarzanie danych.

Innym sposobem na optymalizację jest denormalizacja danych. Polega to na przechowywaniu duplikatów danych w tabeli, co 
pozwala na uniknięcie skomplikowanych złączeń między tabelami i przyspieszenie operacji związanych z pobieraniem 
i zapisywaniem danych.

Ostatecznie, należy pamiętać o wybieraniu odpowiedniego typu danych dla kolumn w tabeli, unikaniu pustych wartości, 
usuwaniu zbędnych kolumn i przestrzeganiu zasad normalizacji bazy danych, aby zapewnić spójność i integralność danych.
"""

"""
Do optymalizacji przechowywania danych o sprzedaży mieszkań w bazie danych można zastosować kilka technik:

    1. Indeksowanie kolumn: 
    Wprowadzenie indeksów na kolumny często wykorzystywane w zapytaniach SQL może znacznie przyspieszyć ich wykonanie. 
    Dzięki indeksowaniu bazy danych nie musi przeszukiwać całej tabeli, aby znaleźć konkretne rekordy.

    2. Normalizacja bazy danych: 
    Normalizacja bazy danych polega na podziale danych na mniejsze, bardziej znormalizowane części, aby uniknąć 
    duplikacji danych. Może to zmniejszyć rozmiar bazy danych i przyspieszyć czas przeszukiwania.

    3. Partycjonowanie tabel: 
    Partycjonowanie tabel polega na podziale tabeli na mniejsze części zwane partycjami. 
    Każda partycja zawiera tylko określone rekordy, co umożliwia szybszy dostęp do danych.

    4. Wykorzystanie pamięci podręcznej: 
    Często wykorzystywane zapytania mogą być przechowywane w pamięci podręcznej bazy danych, aby przyspieszyć czas ich 
    wykonywania.

    5. Użycie odpowiedniego typu danych: 
    Używanie odpowiedniego typu danych dla kolumn może znacznie wpłynąć na wydajność przetwarzania danych. Na przykład 
    używanie typu INTEGER zamiast TEXT w kolumnie zawierającej numery zwiększy wydajność przetwarzania.

    6. Użycie odpowiedniego silnika bazy danych: 
    Wybór odpowiedniego silnika bazy danych może znacznie wpłynąć na wydajność przetwarzania danych. Niektóre silniki 
    są bardziej zoptymalizowane pod kątem szybkiego przeszukiwania dużych baz danych, a inne pod kątem obsługi wielu 
    zapytań równocześnie.

Wprowadzenie tych zmian może pomóc w zwiększeniu wydajności przetwarzania danych i umożliwić szybszy dostęp do danych.
"""
